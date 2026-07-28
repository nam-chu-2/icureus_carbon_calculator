"""Generate london_airfare_high_season.xlsx.

Captures approximate high-season (July) economy round-trip fares from
North American cities to London, England. To add a city, append a dict
to CITIES and re-run:  python build_fares_workbook.py
"""

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

QUERY_DATE = "2026-07-28"
UPLIFT = 0.13  # May shoulder-season advance fare -> July peak estimate

CITIES = [
    {
        "city": "Ottawa, ON, Canada",
        "airport": "YOW (Ottawa Macdonald-Cartier)",
        "drive_note": "Local airport — no drive needed",
        "routing": "1 stop via Montreal (YUL) or Toronto (YYZ); no nonstop to London from YOW",
        "advance_fare_cad": 1175,
        "alternate": "Drive ~2 h to Montreal (YUL): British Airways nonstop ~CA$1,007 RT",
        "notes": "All routings 1-stop; Google Flights flagged queried dates as 'high'",
    },
    {
        "city": "Saguenay, QC, Canada",
        "airport": "YBG (Saguenay-Bagotville)",
        "drive_note": "Local airport — tickets through despite small size",
        "routing": "Air Canada, 1 stop via Montreal (YUL)",
        "advance_fare_cad": 1223,
        "alternate": "Drive ~4.5 h to Montreal (YUL) saves only ~CA$215 — YBG is the practical choice; Quebec City (YQB) ~2 h is the nearest mid-size alternative",
        "notes": "Only ~CA$50 above Ottawa despite being a regional airport",
    },
]

METHOD_LINES = [
    ("Goal", "Approximate high-season (July) economy round-trip airfare from North American cities to London, England (any London airport)."),
    ("Primary source", "Google Flights, priced in CAD, round-trip economy, representative ~2-week travel window, lowest reasonable fare recorded."),
    ("Short-notice confound", "Querying July 2026 travel in late July 2026 is a near-zero-notice booking and overstates a normal high-season fare. July 2027 is beyond the ~11-month airline bookable window, so a clean live July advance quote is impossible."),
    ("Advance-purchase proxy", f"Fares queried {QUERY_DATE} for mid-May 2027 travel — the furthest-out bookable comparable dates, ~10 months of lead time — capturing true advance-purchase pricing with no last-minute premium."),
    ("Seasonal uplift", f"A ~{UPLIFT:.0%} uplift is applied to the May shoulder-season advance fare to estimate the July peak. July estimate = advance fare x {1 + UPLIFT:.2f}, rounded to the nearest CA$10."),
    ("Uncertainty", "Treat all figures as approximate, roughly +/-15%."),
    ("Airport selection rule", "Use the local airport where it tickets through reasonably. Drive up to ~1.5 h to a larger airport if near a smaller one; drive further only if there is no nearby airport. A cheaper alternate hub is noted where one exists."),
    ("Caveat", "On some Canadian routes, summer leisure carriers (e.g., Air Transat) add capacity that can hold July fares down, so July is not always the single priciest month."),
    ("Cross-check", "Sanity-checked against Google Flights price insights and Kayak historical typical-price ranges for the route."),
]

FARE_HEADERS = [
    "City",
    "Departure airport",
    "Drive note",
    "Routing",
    f"Advance RT fare, CAD (mid-May 2027, queried {QUERY_DATE})",
    "Seasonal uplift",
    "Approx July high-season RT, CAD",
    "Cheaper alternate option",
    "Notes",
    "Source",
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")


def july_estimate(advance_fare):
    return round(advance_fare * (1 + UPLIFT) / 10) * 10


def build():
    wb = Workbook()

    ws = wb.active
    ws.title = "Fares"
    ws.append(FARE_HEADERS)
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP

    for c in CITIES:
        ws.append([
            c["city"],
            c["airport"],
            c["drive_note"],
            c["routing"],
            c["advance_fare_cad"],
            UPLIFT,
            july_estimate(c["advance_fare_cad"]),
            c["alternate"],
            c["notes"],
            "Google Flights (CAD)",
        ])

    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = WRAP
        row[4].number_format = '"CA$"#,##0'
        row[5].number_format = "0%"
        row[6].number_format = '"CA$"#,##0'
        row[6].font = Font(bold=True)

    widths = [22, 26, 26, 34, 20, 10, 16, 40, 34, 18]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    ws2 = wb.create_sheet("Method")
    ws2.append(["Topic", "Detail"])
    for cell in ws2[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
    for topic, detail in METHOD_LINES:
        ws2.append([topic, detail])
    for row in ws2.iter_rows(min_row=2):
        row[0].font = Font(bold=True)
        row[0].alignment = Alignment(vertical="top")
        row[1].alignment = WRAP
    ws2.column_dimensions["A"].width = 24
    ws2.column_dimensions["B"].width = 110

    out = "london_airfare_high_season.xlsx"
    wb.save(out)
    print(f"Wrote {out}")


if __name__ == "__main__":
    build()
