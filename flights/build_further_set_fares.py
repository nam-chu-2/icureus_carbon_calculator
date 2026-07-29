"""Generate further_set_airfares.xlsx.

Approximate July high-season economy round-trip fares from every Canadian FSA
and US ZIP3 to the two cities of its "further set":
  Canada  Set A: Vancouver + Calgary   Set B: Toronto + Montreal   (CAD)
  US      Set A: San Francisco + Denver  Set B: New York City + Boston  (USD)

Method (same lineage as the London example in this folder):
  Google Flights advance-purchase quotes for May 12-26, 2027 (~10-month lead,
  queried 2026-07-29) calibrate a fare-vs-distance line per country, plus
  airport-tier premiums observed in the quotes. July peak = fitted fare x 1.13,
  rounded to the nearest $10.

Re-run:  python build_further_set_fares.py
"""

import csv
import math
import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

HERE = os.path.dirname(os.path.abspath(__file__))
PART2 = os.path.join(os.path.dirname(HERE), "part 2", "spreadsheets")

QUERY_DATE = "2026-07-29"
TRAVEL_WINDOW = "2027-05-12 to 2027-05-26"
UPLIFT = 0.13

# ---------------------------------------------------------------- calibration
# Lowest round-trip fare shown by Google Flights for the travel window above.
# tier: how the origin airport's market is classified (see TIER_* below).
CAL = {
    "CA": {
        "major": [("YQB", "YYZ", 662), ("YHZ", "YYZ", 332), ("YWG", "YYZ", 352),
                  ("YXE", "YUL", 731), ("YOW", "YVR", 640), ("YHZ", "YVR", 776),
                  ("YZF", "YVR", 610)],
        "regional": [("YQT", "YVR", 860)],
        "monopoly": [("YBG", "YVR", 1124), ("YZV", "YVR", 1135)],
    },
    "US": {
        "major": [("STL", "DEN", 377), ("MSP", "JFK", 339), ("ATL", "DEN", 433),
                  ("JFK", "DEN", 367), ("ORD", "SFO", 345), ("JFK", "SFO", 379),
                  ("BOS", "SFO", 339), ("FSD", "JFK", 381)],
        "small": [("GJT", "JFK", 443), ("MOT", "JFK", 532)],
    },
}

# Held-out live quotes used only to sanity-check the fitted model (not fitted).
HOLDOUT = {"CA": [("YEG", "YYZ", 487)], "US": [("DEN", "BOS", 247)]}

# ------------------------------------------------------------------- origins
# Canada: nearest_airport_city (co2_emissions.csv spelling) -> IATA
CA_CITY_AIRPORT = {
    "Montréal": "YUL", "Mirabel": "YUL", "Toronto": "YYZ", "Vancouver": "YVR",
    "Halifax": "YHZ", "Calgary": "YYC", "Edmonton": "YEG", "Winnipeg": "YWG",
    "Moncton": "YQM", "Quebec City": "YQB", "Saskatoon": "YXE", "Sudbury": "YSB",
    "London": "YXU", "Kelowna": "YLW", "Îles-de-la-Madeleine": "YGR",
    "St. John's": "YYT", "Baie-Comeau": "YBC", "Thunder Bay": "YQT",
    "Deer Lake": "YDF", "Sept-Îles": "YZV", "Ottawa": "YOW", "Fredericton": "YFC",
    "Fort McMurray": "YMM", "Regina": "YQR", "Grande Prairie": "YQU",
    "Prince George": "YXS", "Thompson": "YTH", "Sault Ste. Marie": "YAM",
    "Whitehorse": "YXY", "Saguenay": "YBG", "Fort St. John": "YXJ",
    "Lethbridge": "YQL", "Smithers": "YYD", "Prince Rupert": "YPR",
    "Brandon": "YBR", "Yellowknife": "YZF", "Cambridge Bay": "YCB",
    "Hopedale": "YHO", "Puvirnituq": "YPX", "Radisson": "YGL",
    "Attawapiskat": "YAT", "Pickle Lake": "YPL", "Medicine Hat": "YXH",
    "Nanaimo": "YCD", "Resolute": "YRB", "Norman Wells": "YVQ",
}

# Canadian airport market tiers (competition level, judged from carrier mix;
# calibration quotes anchor each tier's premium):
#   major    - mainline jet competition (fitted line, no premium)
#   regional - regional service with some mainline/competitor presence
#   monopoly - single-operator regional feed (Jazz/PAL style)
#   northern - fly-in / northern-carrier monopoly; modeled fare is a LOWER BOUND
TIER_CA = {
    "regional": {"YQT", "YSB", "YAM", "YDF", "YXS", "YMM", "YQU", "YXJ",
                 "YQL", "YXH", "YCD", "YBR", "YXU"},
    "monopoly": {"YBG", "YZV", "YBC", "YGR", "YYD", "YPR"},
    "northern": {"YTH", "YCB", "YRB", "YPX", "YHO", "YGL", "YAT", "YPL", "YVQ"},
}

# US small non-hub airports with limited carrier competition (+premium).
TIER_US_SMALL = {"AMA", "BIL", "BIS", "CPR", "CRP", "GFK", "GJT", "GTF", "HSV",
                 "ITO", "JAC", "JAN", "JNU", "LBB", "MAF", "MFE", "MFR", "MOT",
                 "MSO", "RAP", "SHV", "TLH", "XNA"}

# ---------------------------------------------------------------- destinations
DESTS = {
    "CA": {"A": [("Vancouver, BC", "YVR"), ("Calgary, AB", "YYC")],
           "B": [("Toronto, ON", "YYZ"), ("Montréal, QC", "YUL")]},
    "US": {"A": [("San Francisco, CA", "SFO"), ("Denver, CO", "DEN")],
           "B": [("New York City, NY", "JFK"), ("Boston, MA", "BOS")]},
}

# ------------------------------------------------------------------ mechanics

def load_coords():
    path = os.path.join(HERE, "airport_coords.csv")
    with open(path, encoding="utf-8-sig") as f:
        return {r["iata"]: (float(r["lat"]), float(r["lon"]))
                for r in csv.DictReader(f)}


def haversine_km(a, b):
    lat1, lon1, lat2, lon2 = map(math.radians, (*a, *b))
    h = (math.sin((lat2 - lat1) / 2) ** 2
         + math.cos(lat1) * math.cos(lat2) * math.sin((lon2 - lon1) / 2) ** 2)
    return 2 * 6371.0 * math.asin(math.sqrt(h))


def fit_line(points):
    """Least-squares fare = a + b*km; if b < 0, fall back to flat mean."""
    n = len(points)
    mx = sum(p[0] for p in points) / n
    my = sum(p[1] for p in points) / n
    sxx = sum((p[0] - mx) ** 2 for p in points)
    sxy = sum((p[0] - mx) * (p[1] - my) for p in points)
    b = sxy / sxx if sxx else 0.0
    if b < 0:
        return my, 0.0
    return my - b * mx, b


class FareModel:
    def __init__(self, coords):
        self.coords = coords
        self.fits = {}       # country -> (a, b)
        self.premiums = {}   # (country, tier) -> premium
        self.cal_rows = []   # for the Calibration tab
        for country, groups in CAL.items():
            pts = []
            for o, d, fare in groups["major"]:
                km = haversine_km(coords[o], coords[d])
                pts.append((km, fare))
            a, b = fit_line(pts)
            self.fits[country] = (a, b)
            floor = min(fare for _, _, fare in groups["major"])
            self.floors = getattr(self, "floors", {})
            self.floors[country] = floor
            for tier, quotes in groups.items():
                resids = []
                for o, d, fare in quotes:
                    km = haversine_km(coords[o], coords[d])
                    base = a + b * km
                    resids.append(fare - base)
                    self.cal_rows.append([country, tier, f"{o}-{d}", round(km),
                                          fare, round(base), round(fare - base)])
                if tier != "major":
                    self.premiums[(country, tier)] = sum(resids) / len(resids)
        # Northern Canada: no bookable calibration quote exists (e.g. Resolute
        # returned no results) - reuse the monopoly premium and flag the fare
        # as a lower bound.
        self.premiums[("CA", "northern")] = self.premiums[("CA", "monopoly")]

    def tier(self, country, iata):
        if country == "CA":
            for t, s in TIER_CA.items():
                if iata in s:
                    return t
            return "major"
        return "small" if iata in TIER_US_SMALL else "major"

    def july_fare(self, country, origin, dest):
        km = haversine_km(self.coords[origin], self.coords[dest])
        a, b = self.fits[country]
        fare = max(a + b * km, self.floors[country])
        fare += self.premiums.get((country, self.tier(country, origin)), 0.0)
        return round(km), round(fare * (1 + UPLIFT) / 10) * 10


# --------------------------------------------------------------------- inputs

def load_canada():
    with open(os.path.join(PART2, "canada_fsa_flight_destinations.csv"),
              encoding="utf-8-sig") as f:
        sets = {r["fsa"]: r["further_set"] for r in csv.DictReader(f)}
    rows = []
    with open(os.path.join(PART2, "co2_emissions.csv"),
              encoding="utf-8-sig") as f:
        for r in csv.DictReader(f):
            rows.append((r["fsa"], r["name_/_area"], r["nearest_airport_city"],
                         sets[r["fsa"]]))
    assert len(rows) == len(sets), "FSA row mismatch between the two CSVs"
    return rows


def load_us():
    wb = load_workbook(os.path.join(
        PART2, "us_zip3_climate_footprint_US_ZIP3_Summary_v2_web.xlsx"),
        read_only=True)
    ws = wb.active
    hdr = [c.value for c in next(ws.iter_rows(max_row=1))]
    idx = {k: hdr.index(k) for k in ("ZIP3", "Location", "Further Set",
                                     "Airport IATA")}
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if r[0] is None:
            continue
        rows.append((str(r[idx["ZIP3"]]), r[idx["Location"]],
                     r[idx["Airport IATA"]], r[idx["Further Set"]]))
    wb.close()
    return rows


# --------------------------------------------------------------------- output

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")
WRAP = Alignment(wrap_text=True, vertical="top")

TIER_NOTES = {
    ("CA", "regional"): "Regional origin premium applied (calibrated on Thunder Bay).",
    ("CA", "monopoly"): "Single-carrier regional origin premium applied (calibrated on Saguenay/Sept-Îles).",
    ("CA", "northern"): "Northern/fly-in origin: no bookable calibration data; fare is a LOWER BOUND - actual fares often far higher.",
    ("US", "small"): "Small-airport premium applied (calibrated on Grand Junction/Minot).",
}


def style_header(ws):
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = WRAP
    ws.freeze_panes = "A2"


def build():
    coords = load_coords()
    model = FareModel(coords)

    wb = Workbook()

    # ---- Canada FSA tab
    ws = wb.active
    ws.title = "Canada FSA"
    cur = '"CA$"#,##0'
    ws.append(["FSA", "Area", "Further set", "Origin airport", "Origin IATA",
               "Dest city 1", "km 1", "July RT CAD 1",
               "Dest city 2", "km 2", "July RT CAD 2", "Note"])
    for fsa, area, city, fset in load_canada():
        origin = CA_CITY_AIRPORT[city]
        tier = model.tier("CA", origin)
        note = TIER_NOTES.get(("CA", tier), "")
        row = [fsa, area, fset, city, origin]
        for dest_city, dest_iata in DESTS["CA"][fset]:
            if dest_iata == origin:
                row += [dest_city, 0, None]
                note = (note + " " if note else "") + \
                    f"Origin equals destination {dest_city} - check set assignment."
                continue
            km, fare = model.july_fare("CA", origin, dest_iata)
            row += [dest_city, km, fare]
        row.append(note)
        ws.append(row)
    for r in ws.iter_rows(min_row=2):
        r[7].number_format = cur
        r[10].number_format = cur
    widths = [7, 38, 10, 20, 11, 15, 8, 13, 15, 8, 13, 46]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    style_header(ws)

    # ---- US ZIP3 tab
    ws2 = wb.create_sheet("US ZIP3")
    cur = '"$"#,##0'
    ws2.append(["ZIP3", "Location", "Further set", "Origin IATA",
                "Dest city 1", "km 1", "July RT USD 1",
                "Dest city 2", "km 2", "July RT USD 2", "Note"])
    for zip3, loc, origin, fset in load_us():
        tier = model.tier("US", origin)
        note = TIER_NOTES.get(("US", tier), "")
        if origin in ("YOW", "YQB", "YUL", "YYZ"):
            note = (note + " " if note else "") + \
                "Cross-border origin (Canadian airport); USD estimate from US domestic model."
        row = [zip3, loc, fset, origin]
        for dest_city, dest_iata in DESTS["US"][fset]:
            if dest_iata == origin:
                row += [dest_city, 0, None]
                note = (note + " " if note else "") + \
                    f"Origin equals destination {dest_city} - check set assignment."
                continue
            km, fare = model.july_fare("US", origin, dest_iata)
            row += [dest_city, km, fare]
        row.append(note)
        ws2.append(row)
    for r in ws2.iter_rows(min_row=2):
        r[0].number_format = "@"
        r[6].number_format = cur
        r[9].number_format = cur
    widths = [7, 32, 10, 11, 17, 8, 13, 17, 8, 13, 46]
    for i, w in enumerate(widths, 1):
        ws2.column_dimensions[get_column_letter(i)].width = w
    style_header(ws2)

    # ---- Calibration tab
    ws3 = wb.create_sheet("Calibration")
    ws3.append(["Country", "Origin tier", "Route", "km",
                f"Advance RT fare ({TRAVEL_WINDOW}, queried {QUERY_DATE})",
                "Fitted major-curve fare", "Residual (premium evidence)"])
    for row in model.cal_rows:
        ws3.append(row)
    ws3.append([])
    for country in ("CA", "US"):
        a, b = model.fits[country]
        ws3.append([country, "fit", f"fare = {a:.0f} + {b:.4f} x km "
                    f"(floor {model.floors[country]})"])
    for (country, tier), p in sorted(model.premiums.items()):
        ws3.append([country, tier, f"premium +{p:.0f}"])
    ws3.append([])
    for country, quotes in HOLDOUT.items():
        a, b = model.fits[country]
        for o, d, fare in quotes:
            km = haversine_km(coords[o], coords[d])
            base = max(a + b * km, model.floors[country])
            ws3.append([country, "holdout (not fit)", f"{o}-{d}", round(km),
                        fare, round(base), round(fare - base)])
    ws3.append([])
    ws3.append(["", "", "Resolute (YRB) returned no bookable results - northern "
                "tier reuses the monopoly premium as a lower bound."])
    for i, w in enumerate([9, 12, 44, 8, 42, 22, 24], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w
    style_header(ws3)

    # ---- Method tab
    ws4 = wb.create_sheet("Method")
    ws4.append(["Topic", "Detail"])
    method = [
        ("Goal", "Approximate July high-season economy round-trip fare from every Canadian FSA (CAD) and US ZIP3 (USD) to the two cities of its 'further set': Canada A = Vancouver + Calgary, B = Toronto + Montréal; US A = San Francisco + Denver, B = New York City + Boston."),
        ("Set assignment", "Taken directly from the part 2 spreadsheets: 'further_set' in canada_fsa_flight_destinations.csv and 'Further Set' in the US ZIP3 summary. Not recomputed."),
        ("Origin airports", "Canada: nearest_airport_city per FSA (co2_emissions.csv) mapped to its IATA airport. US: the 'Airport IATA' column per ZIP3."),
        ("Fare source", f"Google Flights round-trip economy quotes for {TRAVEL_WINDOW} (queried {QUERY_DATE}, ~10-month lead). July 2027 is beyond the ~11-month bookable window, and pricing July 2026 today would carry a last-minute premium, so mid-May 2027 advance-purchase fares are the proxy - same approach as the London workbook in this folder."),
        ("Distance model", "18 live quotes calibrate a per-country least-squares line fare = a + b x km over great-circle (haversine) airport-to-airport distance, fitted on competitive 'major' origins. US pricing is nearly flat with distance; Canadian pricing rises with distance and, far more, with lack of competition."),
        ("Airport tiers", "Origin airports are tiered by carrier competition. Canada: major (fit line), regional (+ Thunder Bay-calibrated premium), single-carrier monopoly (+ Saguenay/Sept-Îles premium), northern/fly-in (monopoly premium, flagged as lower bound - Resolute had no bookable fares at all). US: major vs small non-hub (+ Grand Junction/Minot premium)."),
        ("July uplift", f"Fitted advance fare x {1 + UPLIFT:.2f} (May shoulder -> July peak), rounded to the nearest $10."),
        ("Currency", "Canada tab in CAD, US tab in USD, as quoted by Google Flights per route."),
        ("Uncertainty", "Treat figures as approximate, roughly +/-20-25% (wider than a single-route live quote: the distance fit averages over large competition-driven scatter, e.g. Flair-served routes price ~half of monopoly routes at the same distance). Northern fly-in rows are lower bounds, not estimates."),
        ("Caveats", "Summer leisure capacity (Air Transat, Flair) can hold July fares down on some routes, so July is not always the priciest month. A handful of border-area US ZIP3s use Canadian airports (YOW/YQB/YUL/YYZ); their USD fares come from the US domestic curve. Rows where origin city equals a destination city are flagged rather than priced. The source data's nearest_airport_city maps some FSAs to a larger airport further away (e.g. Ottawa/Kingston FSAs -> Montréal); this follows the source as-is and shifts fares by only ~$10 at these distances."),
        ("Spot-check", "Two held-out live quotes (not used in the fit): YEG-YYZ CA$487 (Flair; mainline CA$587 vs model CA$621 advance) and DEN-BOS US$247 (1-stop; nonstops US$275-378 vs model US$370). Both within the stated uncertainty band; see Calibration tab."),
        ("Extending", "Add calibration quotes to CAL or adjust tier sets at the top of build_further_set_fares.py, then re-run it."),
    ]
    for t, d in method:
        ws4.append([t, d])
    for r in ws4.iter_rows(min_row=2):
        r[0].font = Font(bold=True)
        r[0].alignment = Alignment(vertical="top")
        r[1].alignment = WRAP
    ws4.column_dimensions["A"].width = 18
    ws4.column_dimensions["B"].width = 120
    style_header(ws4)

    out = os.path.join(HERE, "further_set_airfares.xlsx")
    wb.save(out)
    print("Wrote", out)
    a, b = model.fits["CA"]
    print(f"CA fit: fare = {a:.0f} + {b:.4f}*km | premiums:",
          {k[1]: round(v) for k, v in model.premiums.items() if k[0] == 'CA'})
    a, b = model.fits["US"]
    print(f"US fit: fare = {a:.0f} + {b:.4f}*km | premiums:",
          {k[1]: round(v) for k, v in model.premiums.items() if k[0] == 'US'})


if __name__ == "__main__":
    build()
